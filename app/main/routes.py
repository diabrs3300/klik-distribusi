import io
import logging
from functools import wraps
from flask import render_template, request, redirect, url_for, flash, send_file, abort, jsonify, session
from flask_login import login_required, current_user

from app.main import main
from app.brs_cols import BRS_CONFIG
from app.services.sheets import (
    get_docs, get_klik_links, clear_docs_cache, clear_klik_cache,
    get_progress, clear_progress, process_ihk_upload,
    process_ekspor_upload, process_impor_upload
)

_log = logging.getLogger(__name__)

@main.context_processor
def inject_brs_config():
    return dict(brs_config=BRS_CONFIG)

def _require_akses(flag_key):
    """Decorator: tolak akses (403) jika user tidak punya flag_key."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if not current_user.akses.get(flag_key, False):
                abort(403)
            return f(*args, **kwargs)
        return wrapped
    return decorator



def _safe_int(val):
    """Konversi nilai ke int dengan aman — menangani string, float, dan NaN."""
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None

def _get_upload_params(form):
    """Extract common upload parameters from form."""
    return {
        'bulan': form.get('bulan', '').strip(),
        'tahun': form.get('tahun', '').strip(),
        'task_id': form.get('task_id', '').strip()
    }

def _validate_basic_upload(bulan, tahun, file, required_ext='.xlsx'):
    """Common validation for upload forms."""
    errors = []
    if not bulan:
        errors.append('Bulan wajib dipilih.')
    if not tahun:
        errors.append('Tahun wajib diisi.')
    elif not tahun.isdigit() or len(tahun) != 4:
        errors.append('Tahun harus 4 digit angka.')
    
    if not file or file.filename == '':
        errors.append('File wajib diunggah.')
    elif not file.filename.lower().endswith(required_ext):
        errors.append(f'File harus berformat {required_ext}')
    
    return errors

def _find_header_row(file, required_cols):
    """
    Mencari baris header di Excel dengan mencocokkan required_cols.
    Mengembalikan index baris (0-indexed). Default 0 jika tidak ditemukan.
    """
    import pandas as pd
    try:
        # Intip 20 baris pertama
        df_peek = pd.read_excel(file, header=None, nrows=20, dtype=str).fillna('')
        file.seek(0) # Reset pointer
        
        # Bersihkan target kolom (strip & lowercase)
        targets = set(str(c).strip().lower() for c in required_cols)
        
        for idx, row in df_peek.iterrows():
            # Ambil semua nilai dalam baris ini yang tidak kosong
            row_set = set(str(val).strip().lower() for val in row.values if str(val).strip())
            
            # Jika semua kolom wajib ada di baris ini
            if targets.issubset(row_set):
                _log.info(f"Header ditemukan di baris ke-{idx+1}")
                return idx
        
        _log.warning("Header tidak ditemukan secara otomatis, menggunakan baris default (0/1).")
        return 0
    except Exception as e:
        _log.error(f"Gagal deteksi header: {e}")
        file.seek(0)
        return 0

def _read_uploaded_file(file, config_key=None, header=None, required_cols=None):
    """Read uploaded Excel or DBF file and return DataFrame."""
    import pandas as pd
    import os
    import tempfile
    
    filename = file.filename.lower()
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Jika header None dan ada required_cols, deteksi otomatis
        if header is None and required_cols:
            header = _find_header_row(file, required_cols)
        elif header is None:
            header = 0
            
        df = pd.read_excel(file, header=header, dtype=str).fillna('')
        df.columns = [str(c).strip() for c in df.columns]
        return df
    
    if filename.endswith('.dbf'):
        from dbfread import DBF
        fd, temp_path = tempfile.mkstemp(suffix='.dbf')
        os.close(fd)
        try:
            file.save(temp_path)
            dbf = DBF(temp_path, load=True, lowernames=False)
            df = pd.DataFrame(iter(dbf))
            df.columns = [c.strip().upper() for c in df.columns]
            return df
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)
    
    return None

def _generate_excel_template(config_key, download_name, title_text=None):
    """Generic helper to generate Excel template based on BRS_CONFIG."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    config = BRS_CONFIG[config_key]
    color = config.get('theme_color', '0D6EFD')
    
    # Try all possible column mapping keys
    cols_dict = (
        config.get('required_cols') or 
        config.get('required_excel_cols') or 
        config.get('excel_cols')
    )
    
    if isinstance(cols_dict, dict):
        # All configs now: SemanticKey → Excel col name, so use .values() for header
        cols = list(cols_dict.values())
    else:
        cols = cols_dict
        
    sample = config.get('template', {}).get('sample', [])
    col_widths = config.get('template', {}).get('col_widths', [15] * len(cols))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Template {config_key.upper()}'

    # Header section
    start_row = 1
    if title_text:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        cell = ws.cell(row=1, column=1, value=title_text)
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=color.replace('#', ''))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24
        
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
        note = ws.cell(row=2, column=1, value='Header kolom WAJIB berada di BARIS KE-3. Isi data mulai baris ke-4.')
        note.font = Font(italic=True, size=10, color='856404')
        note.fill = PatternFill('solid', fgColor='FFF3CD')
        note.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18
        start_row = 3

    # Column headers
    hfill = PatternFill('solid', fgColor=color.replace('#', ''))
    hfont = Font(bold=True, color='FFFFFF', size=10)
    hbord = Border(bottom=Side(style='medium'), right=Side(style='thin', color='CCCCCC'))
    
    # Track columns for specific formatting
    numeric_cols = []
    text_cols = []
    for i, col in enumerate(cols, start=1):
        # Numeric columns that shouldn't show scientific notation (E+)
        if col in ('NK', 'FOB', 'NILAI', 'Nilai', 'IHK', 'TPK', 'RATS'):
            numeric_cols.append(i)
        
        # Identity/Code columns that should preserve leading zeros
        if col in ('Kode Kota', 'Kode Komoditas', 'Kode', 'Kd.Kota', 'KODE_HS', 'KODE BTKI', 'HS Code', 'HS'):
            text_cols.append(i)
        
        cell = ws.cell(row=start_row, column=i, value=col)
        cell.font, cell.fill, cell.border = hfont, hfill, hbord
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1] if i-1 < len(col_widths) else 15
    
    ws.row_dimensions[start_row].height = 30

    # Apply formatting to data rows (Sample + Next 100 rows for easier manual entry)
    for r in range(start_row + 1, start_row + 102):
        for i in numeric_cols:
            ws.cell(row=r, column=i).number_format = '0.00'
        for i in text_cols:
            ws.cell(row=r, column=i).number_format = '@'

    # Insert sample data
    if sample:
        data_row = start_row + 1
        for i, val in enumerate(sample, start=1):
            cell = ws.cell(row=data_row, column=i, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.freeze_panes = f'A{start_row + 1}'
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=download_name,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@main.route('/')
@main.route('/index')
def index():
    klik_groups = get_klik_links()
    return render_template('index.html', title='Klik Distribusi', klik_groups=klik_groups)


@main.route('/dia-brs/klik-refresh')
def klik_refresh():
    """Hapus cache Klik Distribusi lalu redirect ke halaman utama."""
    clear_klik_cache()
    flash('Klik Distribusi telah di-refresh — Daftar link berhasil diperbarui.', 'success')
    return redirect(url_for('main.index'))


@main.route('/dia-brs/docs-refresh')
@login_required
def docs_refresh():
    """Hapus cache semua BRS docs lalu redirect kembali ke halaman asal."""
    clear_docs_cache()
    flash('Dokumen-dokumen BRS telah di-refresh — Daftar dokumen berhasil diperbarui.', 'success')
    referrer = request.referrer
    if referrer:
        return redirect(referrer)
    return redirect(url_for('main.dia_brs'))


@main.route('/dia-brs/dashboard')
@login_required
def dashboard():
    return redirect(url_for('main.dia_brs'))


@main.route('/dia-brs')
@login_required
def dia_brs():
    return render_template('dashboard.html', title='DIA BRS — Dashboard')


@main.route('/dia-brs/developer')
@login_required
def developer():
    return render_template('developer.html', title='Tim Pengembang — DIA BRS')



@main.route('/dia-brs/ihk')
@login_required
@_require_akses('akses_ihk')
def brs_ihk():
    docs = get_docs('IHK', fallback=[])
    return render_template('brs/ihk.html', title='BRS IHK-Inflasi', docs=docs)


@main.route('/dia-brs/ntp')
@login_required
@_require_akses('akses_ntp')
def brs_ntp():
    docs = get_docs('NTP', fallback=[])
    return render_template('brs/ntp.html', title='BRS NTP', docs=docs)


@main.route('/dia-brs/transportasi')
@login_required
@_require_akses('akses_transportasi')
def brs_transportasi():
    docs = get_docs('Transportasi', fallback=[])
    return render_template('brs/transportasi.html', title='BRS Transportasi', docs=docs)


@main.route('/dia-brs/ekspor-impor')
@login_required
@_require_akses('akses_ekspor_impor')
def brs_ekspor_impor():
    docs = get_docs('Ekspor Impor', fallback=[])
    return render_template('brs/ekspor_impor.html', title='BRS Ekspor Impor', docs=docs)


@main.route('/dia-brs/pariwisata')
@login_required
@_require_akses('akses_pariwisata')
def brs_pariwisata():
    docs = get_docs('Pariwisata', fallback=[])
    return render_template('brs/pariwisata.html', title='BRS Pariwisata', docs=docs)




# ─── Upload IHK ───────────────────────────────────────────────────────────────

# ─── Progress Tracker API ─────────────────────────────────────────────────────

@main.route('/dia-brs/upload-progress/<task_id>')
@login_required
def upload_progress(task_id):
    return jsonify(get_progress(task_id))

def _handle_upload_error(msg, template, title, **kwargs):
    """Helper to handle upload errors consistently for regular and AJAX requests."""
    flash(msg, 'danger')
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'status': 'error', 'message': msg}), 400
    
    # Ensure docs are passed back for templates that use them
    if 'docs' not in kwargs:
        from app.services.sheets import get_docs
        if 'ihk' in template:
            kwargs['docs'] = get_docs('IHK', fallback=[])
        elif 'ekspor_impor' in template:
            kwargs['docs'] = get_docs('Ekspor Impor', fallback=[])
            
    return render_template(template, title=title, **kwargs), 400

@main.route('/dia-brs/ihk/upload', methods=['GET', 'POST'])
@login_required
@_require_akses('akses_ihk')
def upload_ihk():
    if request.method == 'GET':
        docs = get_docs('IHK', fallback=[])
        return render_template('brs/upload_ihk.html', title='Upload Excel IHK/Inflasi', docs=docs)

    params = _get_upload_params(request.form)
    file = request.files.get('file')
    
    errors = _validate_basic_upload(params['bulan'], params['tahun'], file)
    if errors:
        return _handle_upload_error(". ".join(errors), 'brs/upload_ihk.html', 'Upload Excel IHK/Inflasi')

    # ── Baca Excel ────────────────────────────────────────────────────────────
    try:
        df = _read_uploaded_file(file, required_cols=BRS_CONFIG['ihk']['required_cols'].values())
    except Exception as e:
        return _handle_upload_error(f'Gagal membaca file Excel: {e}', 'brs/upload_ihk.html', 'Upload Excel IHK/Inflasi')

    # Build cols_map: config default + user overrides
    cols_map = dict(BRS_CONFIG['ihk']['required_cols'])
    for sem_key in list(cols_map.keys()):
        user_col = request.form.get(f'col_map[{sem_key}]', '').strip()
        if user_col:
            cols_map[sem_key] = user_col

    # Case-insensitive robust rename: map actual columns (upper-stripped) to SemanticKey
    rename_map = {v.strip().upper(): k for k, v in cols_map.items()}
    df = df.rename(columns={c: rename_map.get(c.upper(), c) for c in df.columns})

    # ── Validasi kolom SemanticKey ────────────────────────────────────────────
    missing_cols = [k for k in cols_map if k not in df.columns]
    if missing_cols:
        excel_names = [cols_map[k] for k in missing_cols]
        msg = f'Kolom Excel tidak lengkap. Kolom yang kurang: {", ".join(excel_names)}'
        return _handle_upload_error(msg, 'brs/upload_ihk.html', 'Upload Excel IHK/Inflasi')

    if df.empty:
        flash('File Excel tidak memiliki data.', 'warning')
        return render_template('brs/upload_ihk.html', title='Upload Excel IHK-Inflasi'), 400

    # ── Validasi & Filter Bulan/Tahun ─────────────────────────────────────────
    bulan_int, tahun_str = int(params['bulan']), params['tahun']
    mask = df.apply(
        lambda r: str(r.get('KeyTahun', '')).strip() == tahun_str and
                  _safe_int(r.get('KeyBulan', '')) == bulan_int,
        axis=1
    )
    
    matched_df = df[mask]
    if matched_df.empty:
        msg = f'Tidak ada data untuk Bulan {params["bulan"]} Tahun {params["tahun"]} di file Excel.'
        return _handle_upload_error(msg, 'brs/upload_ihk.html', 'Upload Excel IHK-Inflasi')

    filter_info = f'Hanya {len(matched_df)} dari {len(df)} baris yang sesuai — baris lain dilewati.' if len(matched_df) < len(df) else None
    col_name = tahun_str[2:] + params['bulan'].zfill(2)

    from app.services.sheets import set_progress
    if params['task_id']:
        set_progress(params['task_id'], 5, "Menyiapkan data untuk upload...")

    # ── Proses ke Google Sheets ───────────────────────────────────────────────
    try:
        from app.services.sheets import process_ihk_upload, clear_progress
        stats = process_ihk_upload(matched_df, col_name, task_id=params['task_id'])
        if stats['inserted'] == 0 and stats['updated'] == 0:
            flash('File berhasil dibaca, namun tidak ada baris data yang valid untuk di-upload.', 'warning')
    except FileNotFoundError:
        msg = 'File credentials.json tidak ditemukan. Letakkan service account key di root project.'
        return _handle_upload_error(msg, 'brs/upload_ihk.html', 'Upload Excel IHK/Inflasi')
    except Exception as e:
        return _handle_upload_error(f'Gagal meng-upload ke Google Sheets: {e}', 'brs/upload_ihk.html', 'Upload Excel IHK/Inflasi')

    if params['task_id']: clear_progress(params['task_id'])
    
    # AJAX support for smooth reload + form reset
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        session['upload_stats'] = stats
        session['upload_col_name'] = col_name
        session['upload_filter_info'] = filter_info
        return jsonify({'status': 'success'})

    return render_template('brs/upload_ihk.html', title='Upload Excel IHK-Inflasi', stats=stats, col_name=col_name, filter_info=filter_info)


# ── Download Template Excel ────────────────────────────────────────────────────

@main.route('/dia-brs/ihk/template')
@login_required
@_require_akses('akses_ihk')
def download_template_ihk():
    """Generate dan kirim file template Excel IHK-Inflasi."""
    return _generate_excel_template(
        config_key='ihk',
        download_name='template_ihk_inflasi.xlsx',
        title_text='Template Upload Data IHK-Inflasi — DIA BRS'
    )


# ─── Upload Ekspor-Impor ───────────────────────────────────────────────────────


@main.route('/dia-brs/ekspor-impor/upload', methods=['GET', 'POST'])
@login_required
@_require_akses('akses_ekspor_impor')
def upload_ekspor_impor():
    if request.method == 'GET':
        docs = get_docs('Ekspor Impor', fallback=[])
        return render_template('brs/upload_ekspor_impor.html', title='Upload Ekspor Impor', docs=docs)

    form_type = request.form.get('form_type', '').strip()  # 'ekspor' atau 'impor'
    params = _get_upload_params(request.form)
    file = request.files.get('file')

    if form_type not in ('ekspor', 'impor'):
        return _handle_upload_error('Tipe form tidak valid.', 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor')

    # ── Validasi file ────────────────────────────────────────────────────────
    if not file or file.filename == '':
        return _handle_upload_error('File wajib diunggah.', 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)

    is_excel = file.filename.lower().endswith(('.xlsx', '.xls'))
    if not is_excel and not file.filename.lower().endswith('.dbf'):
        return _handle_upload_error('File harus berformat .dbf atau .xlsx/.xls', 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)

    # ── Baca File ────────────────────────────────────────────────────────────
    try:
        df = _read_uploaded_file(file)
        _config = BRS_CONFIG[form_type]

        if is_excel:
            # Build cols_map: start from config defaults, override with user-supplied col names
            cols_map = dict(_config['required_excel_cols'])
            for sem_key in list(cols_map.keys()):
                user_col = request.form.get(f'col_map[{sem_key}]', '').strip()
                if user_col:
                    cols_map[sem_key] = user_col
            rename_map = {v.strip().upper(): k for k, v in cols_map.items()}
        else:
            # Build cols_map: start from config defaults, override with user-supplied col names
            cols_map = dict(_config['required_dbf_cols'])
            for sem_key in list(cols_map.keys()):
                user_col = request.form.get(f'col_map[{sem_key}]', '').strip()
                if user_col:
                    cols_map[sem_key] = user_col
            rename_map = {v.strip().upper(): k for k, v in cols_map.items()}

        # Robust case-insensitive rename
        df = df.rename(columns={c: rename_map.get(c.upper(), c) for c in df.columns})

        if form_type == 'impor' and not is_excel:
            if not params['bulan'] or not params['tahun']:
                return _handle_upload_error('Bulan dan Tahun wajib diisi untuk upload DBF Impor.', 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)
            df['KeyTahun'], df['KeyBulan'] = params['tahun'], params['bulan'].zfill(2)
            
            # User can override the value column name for impor DBF too
            default_val_col = f"N{params['bulan'].zfill(2)}{params['tahun'][-2:]}"
            user_val_col = request.form.get('col_map[KeyNilai]', '').strip().upper() or default_val_col.upper()
            # Find match in actual df columns (case-insensitive)
            matched_val_col = next((c for c in df.columns if c.upper() == user_val_col), None)
            if not matched_val_col:
                msg = f'Kolom nilai DBF Impor tidak ditemukan: "{user_val_col}". Pastikan nama kolom sudah benar.'
                return _handle_upload_error(msg, 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)
            df = df.rename(columns={matched_val_col: 'KeyNilai'})
            cols_map['KeyNilai'] = matched_val_col

    except Exception as e:
        return _handle_upload_error(f'Gagal membaca file: {e}', 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)

    # ── Validasi & Upload ────────────────────────────────────────────────────
    missing_semantic = [c for c in _config['required_dbf_cols'].keys() if c not in df.columns]
    if missing_semantic:
        missing_real = [cols_map.get(c, c) for c in missing_semantic]
        msg = f'Kolom tidak lengkap. Kurang: {", ".join(missing_real)}'
        return _handle_upload_error(msg, 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)

    if df.empty:
        return _handle_upload_error('File tidak memiliki data.', 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)

    from app.services.sheets import set_progress
    if params['task_id']:
        set_progress(params['task_id'], 5, "Menyiapkan data untuk upload...")

    try:
        if form_type == 'ekspor': stats = process_ekspor_upload(df, task_id=params['task_id'])
        else: stats = process_impor_upload(df, task_id=params['task_id'])
    except Exception as e:
        return _handle_upload_error(f'Gagal upload ke Google Sheets: {e}', 'brs/upload_ekspor_impor.html', 'Upload Ekspor-Impor', active_tab=form_type)
        
    if params['task_id']: clear_progress(params['task_id'])
    
    # AJAX support for smooth reload + form reset
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        session['upload_stats'] = stats
        session['upload_last_type'] = form_type.capitalize()
        session['upload_last_rows'] = len(df)
        session['active_tab'] = form_type
        return jsonify({'status': 'success'})

    return render_template('brs/upload_ekspor_impor.html', title='Upload Ekspor-Impor', active_tab=form_type, last_upload={'type': form_type.capitalize(), 'rows': len(df)}, stats=stats)


# ── Download Template Ekspor-Impor ─────────────────────────────────────────────

@main.route('/dia-brs/ekspor-impor/template-ekspor')
@login_required
@_require_akses('akses_ekspor_impor')
def download_template_ekspor():
    return _generate_excel_template('ekspor', 'template_ekspor.xlsx')

@main.route('/dia-brs/ekspor-impor/template-impor')
@login_required
@_require_akses('akses_ekspor_impor')
def download_template_impor():
    return _generate_excel_template('impor', 'template_impor.xlsx')


# ─── Upload NTP ───────────────────────────────────────────────────────────────

@main.route('/dia-brs/ntp/upload', methods=['GET', 'POST'])
@login_required
@_require_akses('akses_ntp')
def upload_ntp():
    if request.method == 'GET':
        docs = get_docs('NTP', fallback=[])
        return render_template('brs/upload_ntp.html', title='Upload Excel NTP', docs=docs)

    params = _get_upload_params(request.form)
    file = request.files.get('file')
    
    errors = _validate_basic_upload(params['bulan'], params['tahun'], file)
    if errors:
        return _handle_upload_error(". ".join(errors), 'brs/upload_ntp.html', 'Upload Excel NTP')

    try:
        df = _read_uploaded_file(file, required_cols=BRS_CONFIG['ntp']['required_cols'].values())
    except Exception as e:
        return _handle_upload_error(f'Gagal membaca file Excel: {e}', 'brs/upload_ntp.html', 'Upload Excel NTP')

    # Build cols_map: config default + user overrides
    cols_map = dict(BRS_CONFIG['ntp']['required_cols'])
    for sem_key in list(cols_map.keys()):
        user_col = request.form.get(f'col_map[{sem_key}]', '').strip()
        if user_col:
            cols_map[sem_key] = user_col

    # Case-insensitive robust rename
    rename_map = {v.strip().upper(): k for k, v in cols_map.items()}
    df = df.rename(columns={c: rename_map.get(c.upper(), c) for c in df.columns})

    missing = [c for c in BRS_CONFIG['ntp']['required_cols'] if c not in df.columns]
    if missing:
        flash(f'Kolom Excel tidak lengkap. Kurang: {", ".join(missing)}', 'danger')
        return render_template('brs/upload_ntp.html', title='Upload Excel NTP')

    if df.empty:
        flash('File Excel tidak memiliki data.', 'warning')
        return render_template('brs/upload_ntp.html', title='Upload Excel NTP')

    # AJAX support for smooth reload + form reset
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        session['upload_stats'] = {'inserted': 0, 'updated': 0} # Placeholder
        return jsonify({'status': 'success'})

    from app.services.sheets import set_progress
    if params['task_id']:
        set_progress(params['task_id'], 100, "Validasi selesai (Fitur GSheets segera hadir)")

    flash('File berhasil dibaca. Fitur upload NTP ke Google Sheets sedang dalam pengembangan.', 'info')
    docs = get_docs('NTP', fallback=[])
    return render_template('brs/upload_ntp.html', title='Upload Excel NTP', docs=docs)


@main.route('/dia-brs/ntp/template')
@login_required
@_require_akses('akses_ntp')
def download_template_ntp():
    return _generate_excel_template('ntp', 'template_ntp.xlsx', 'Template Upload Data NTP — DIA BRS')


# ─── Upload Pariwisata ────────────────────────────────────────────────────────

@main.route('/dia-brs/pariwisata/upload', methods=['GET', 'POST'])
@login_required
@_require_akses('akses_pariwisata')
def upload_pariwisata():
    if request.method == 'GET':
        docs = get_docs('Pariwisata', fallback=[])
        return render_template('brs/upload_pariwisata.html', title='Upload Excel Pariwisata', docs=docs)

    params = _get_upload_params(request.form)
    file = request.files.get('file')
    
    errors = _validate_basic_upload(params['bulan'], params['tahun'], file)
    if errors:
        return _handle_upload_error(". ".join(errors), 'brs/upload_pariwisata.html', 'Upload Excel Pariwisata')

    try:
        df = _read_uploaded_file(file, required_cols=BRS_CONFIG['pariwisata']['required_cols'].values())
    except Exception as e:
        return _handle_upload_error(f'Gagal membaca file Excel: {e}', 'brs/upload_pariwisata.html', 'Upload Excel Pariwisata')

    # Build cols_map
    cols_map = dict(BRS_CONFIG['pariwisata']['required_cols'])
    for sem_key in list(cols_map.keys()):
        user_col = request.form.get(f'col_map[{sem_key}]', '').strip()
        if user_col:
            cols_map[sem_key] = user_col

    rename_map = {v.strip().upper(): k for k, v in cols_map.items()}
    df = df.rename(columns={c: rename_map.get(c.upper(), c) for c in df.columns})

    missing = [c for c in BRS_CONFIG['pariwisata']['required_cols'] if c not in df.columns]
    if missing:
        msg = f'Kolom Excel tidak lengkap. Kurang: {", ".join(missing)}'
        return _handle_upload_error(msg, 'brs/upload_pariwisata.html', 'Upload Excel Pariwisata')

    if df.empty:
        return _handle_upload_error('File Excel tidak memiliki data.', 'brs/upload_pariwisata.html', 'Upload Excel Pariwisata')

    # AJAX support for smooth reload + form reset
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        session['upload_stats'] = {'inserted': 0, 'updated': 0} # Placeholder
        return jsonify({'status': 'success'})

    from app.services.sheets import set_progress
    if params['task_id']:
        set_progress(params['task_id'], 100, "Validasi selesai (Fitur GSheets segera hadir)")

    flash('File berhasil dibaca. Fitur upload Pariwisata ke Google Sheets sedang dalam pengembangan.', 'info')
    docs = get_docs('Pariwisata', fallback=[])
    return render_template('brs/upload_pariwisata.html', title='Upload Excel Pariwisata', docs=docs)


@main.route('/dia-brs/pariwisata/template')
@login_required
@_require_akses('akses_pariwisata')
def download_template_pariwisata():
    return _generate_excel_template('pariwisata', 'template_pariwisata.xlsx', 'Template Upload Data Pariwisata — DIA BRS')


# ─── Upload Transportasi ──────────────────────────────────────────────────────

@main.route('/dia-brs/transportasi/upload', methods=['GET', 'POST'])
@login_required
@_require_akses('akses_transportasi')
def upload_transportasi():
    if request.method == 'GET':
        docs = get_docs('Transportasi', fallback=[])
        return render_template('brs/upload_transportasi.html', title='Upload Excel Transportasi', docs=docs)

    params = _get_upload_params(request.form)
    file = request.files.get('file')
    
    errors = _validate_basic_upload(params['bulan'], params['tahun'], file)
    if errors:
        for e in errors: flash(e, 'danger')
        return render_template('brs/upload_transportasi.html', title='Upload Excel Transportasi')

    try:
        df = _read_uploaded_file(file, required_cols=BRS_CONFIG['transportasi']['required_cols'].values())
    except Exception as e:
        return _handle_upload_error(f'Gagal membaca file Excel: {e}', 'brs/upload_transportasi.html', 'Upload Excel Transportasi')

    # Build cols_map
    cols_map = dict(BRS_CONFIG['transportasi']['required_cols'])
    for sem_key in list(cols_map.keys()):
        user_col = request.form.get(f'col_map[{sem_key}]', '').strip()
        if user_col:
            cols_map[sem_key] = user_col

    rename_map = {v.strip().upper(): k for k, v in cols_map.items()}
    df = df.rename(columns={c: rename_map.get(c.upper(), c) for c in df.columns})

    missing = [c for c in BRS_CONFIG['transportasi']['required_cols'] if c not in df.columns]
    if missing:
        flash(f'Kolom Excel tidak lengkap. Kurang: {", ".join(missing)}', 'danger')
        return render_template('brs/upload_transportasi.html', title='Upload Excel Transportasi')

    if df.empty:
        flash('File Excel tidak memiliki data.', 'warning')
        return render_template('brs/upload_transportasi.html', title='Upload Excel Transportasi')

    # AJAX support for smooth reload + form reset
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        session['upload_stats'] = {'inserted': 0, 'updated': 0} # Placeholder
        return jsonify({'status': 'success'})

    from app.services.sheets import set_progress
    if params['task_id']:
        set_progress(params['task_id'], 100, "Validasi selesai (Fitur GSheets segera hadir)")

    flash('File berhasil dibaca. Fitur upload Transportasi ke Google Sheets sedang dalam pengembangan.', 'info')
    docs = get_docs('Transportasi', fallback=[])
    return render_template('brs/upload_transportasi.html', title='Upload Excel Transportasi', docs=docs)


@main.route('/dia-brs/transportasi/template')
@login_required
@_require_akses('akses_transportasi')
def download_template_transportasi():
    return _generate_excel_template('transportasi', 'template_transportasi.xlsx', 'Template Upload Data Transportasi — DIA BRS')
